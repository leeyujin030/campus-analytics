"""Excel 리포트 생성기 — 5시트 구성"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ── 색상 팔레트 ────────────────────────────────
C = {
    "dark":    "0D1B2A",
    "navy":    "1E3A5F",
    "accent":  "E94560",
    "blue":    "1D6DE5",
    "green":   "16A34A",
    "warn":    "D97706",
    "gray":    "64748B",
    "light":   "F1F5F9",
    "white":   "FFFFFF",
    "border":  "CBD5E1",
    "header_fg": "FFFFFF",
}

FONT_NAME = "Malgun Gothic"  # 한글 폰트 (Windows/macOS 공용)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=10, color="000000", name=FONT_NAME) -> Font:
    return Font(bold=bold, size=size, color=color, name=name)


def _border_thin() -> Border:
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def _border_medium() -> Border:
    s = Side(style="medium", color=C["dark"])
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _write_df(ws, df: pd.DataFrame, start_row: int, header_color: str,
              number_cols: list[str] | None = None, pct_cols: list[str] | None = None):
    """DataFrame을 시트에 쓰고 스타일 적용"""
    number_cols = number_cols or []
    pct_cols = pct_cols or []

    # 헤더 행
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=ci, value=col)
        cell.font = _font(bold=True, color=C["header_fg"])
        cell.fill = _fill(header_color)
        cell.alignment = _align("center")
        cell.border = _border_thin()

    # 데이터 행
    for ri, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
        for ci, val in enumerate(row, 1):
            col_name = df.columns[ci - 1]
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = _border_thin()
            cell.alignment = _align("right" if col_name in number_cols + pct_cols else "left")

            # 숫자 포맷
            if col_name in number_cols and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
            elif col_name in pct_cols and isinstance(val, (int, float)):
                cell.number_format = "0.0"

            # 짝수 행 배경
            if (ri - start_row) % 2 == 0:
                cell.fill = _fill(C["light"])

    return ri  # 마지막 row 번호


def _set_col_widths(ws, widths: dict[int, int]):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell


def _meta_row(ws, row: int, date_label: str, generated_at: str):
    ws.cell(row=row, column=1,
            value=f"NHN Payco 캠퍼스 사업팀  |  기준일: {date_label}  |  생성: {generated_at}  |  캠퍼스 데일리 애널리틱스 v1.0"
            ).font = _font(size=9, color=C["gray"])


# ── 시트 1: 요약 ──────────────────────────────────────────────────────────────
def _sheet_summary(wb: Workbook, kpi: dict, by_campus: pd.DataFrame,
                   by_cat: pd.DataFrame, date_label: str, generated_at: str):
    ws = wb.create_sheet("요약")
    ws.sheet_view.showGridLines = False

    # 제목
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"캠퍼스 데일리 결제 현황 요약  ({date_label})"
    c.font = _font(bold=True, size=14, color=C["header_fg"])
    c.fill = _fill(C["dark"])
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    # 메타
    ws.merge_cells("A2:J2")
    _meta_row(ws, 2, date_label, generated_at)
    ws.row_dimensions[2].height = 16

    # KPI 카드 (row 4)
    ws.row_dimensions[4].height = 18
    kpi_items = [
        ("총 결제금액", f"₩{kpi['총_결제금액']:,}", C["blue"]),
        ("총 주문건수", f"{kpi['총_주문건수']:,}건", C["navy"]),
        ("캠퍼스 수", f"{kpi['캠퍼스_수']:,}개", C["green"]),
        ("가맹점 수", f"{kpi['가맹점_수']:,}개", C["accent"]),
    ]
    col_starts = [1, 3, 5, 7]
    for (label, val, color), cs in zip(kpi_items, col_starts):
        # 라벨
        lc = ws.cell(row=4, column=cs, value=label)
        lc.font = _font(bold=True, size=9, color=C["header_fg"])
        lc.fill = _fill(color)
        lc.alignment = _align("center")
        ws.merge_cells(start_row=4, start_column=cs, end_row=4, end_column=cs + 1)
        # 값
        vc = ws.cell(row=5, column=cs, value=val)
        vc.font = _font(bold=True, size=13, color=color)
        vc.alignment = _align("center")
        ws.merge_cells(start_row=5, start_column=cs, end_row=5, end_column=cs + 1)
    ws.row_dimensions[5].height = 22

    # 결제유형별 소계 (row 7)
    ws.cell(row=7, column=1, value="결제 유형별 소계").font = _font(bold=True, size=10)
    pay_headers = ["유형", "결제금액", "비중(%)"]
    for ci, h in enumerate(pay_headers, 1):
        c = ws.cell(row=8, column=ci, value=h)
        c.font = _font(bold=True, color=C["header_fg"])
        c.fill = _fill(C["navy"])
        c.alignment = _align("center")
        c.border = _border_thin()

    for ri, (pt, data) in enumerate(kpi["결제유형별"].items(), 9):
        ws.cell(row=ri, column=1, value=pt).border = _border_thin()
        val_cell = ws.cell(row=ri, column=2, value=data["금액"])
        val_cell.number_format = "#,##0"
        val_cell.alignment = _align("right")
        val_cell.border = _border_thin()
        pct_cell = ws.cell(row=ri, column=3, value=data["비중"])
        pct_cell.number_format = "0.0"
        pct_cell.alignment = _align("right")
        pct_cell.border = _border_thin()
        if ri % 2 == 0:
            for ci in range(1, 4):
                ws.cell(row=ri, column=ci).fill = _fill(C["light"])

    # 캠퍼스별 TOP 20 (row 7, col 5~)
    top_cols = ["캠퍼스", "전체_주문", "전체_결제", "카드_결제", "식권_결제", "쿠폰_결제", "승차권_결제"]
    top20 = by_campus[top_cols].head(20).copy()
    top20.columns = ["캠퍼스", "주문건수", "결제금액(원)", "카드", "식권", "쿠폰", "승차권"]

    ws.cell(row=7, column=5, value="캠퍼스별 TOP 20 (결제금액 기준)").font = _font(bold=True, size=10)
    num_cols = ["주문건수", "결제금액(원)", "카드", "식권", "쿠폰", "승차권"]
    _write_df(ws, top20, start_row=8, header_color=C["navy"],
              number_cols=num_cols)

    _set_col_widths(ws, {1: 10, 2: 16, 3: 10, 5: 28, 6: 11, 7: 14, 8: 12, 9: 10, 10: 10, 11: 10})
    _freeze(ws, "A3")


# ── 시트 2: 캠퍼스별 ──────────────────────────────────────────────────────────
def _sheet_campus(wb: Workbook, df: pd.DataFrame, date_label: str, generated_at: str):
    ws = wb.create_sheet("캠퍼스별")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"캠퍼스별 결제 현황  ({date_label})"
    c.font = _font(bold=True, size=13, color=C["header_fg"])
    c.fill = _fill(C["dark"])
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:I2")
    _meta_row(ws, 2, date_label, generated_at)

    out = df[["캠퍼스", "전체_주문", "전체_결제", "카드_주문", "카드_결제",
              "식권_주문", "식권_결제", "쿠폰_주문", "쿠폰_결제",
              "승차권_주문", "승차권_결제"]].copy()
    out.columns = ["캠퍼스", "전체_주문", "전체_결제",
                   "카드_주문", "카드_결제", "식권_주문", "식권_결제",
                   "쿠폰_주문", "쿠폰_결제", "승차권_주문", "승차권_결제"]

    num_cols = [c for c in out.columns if c != "캠퍼스"]
    _write_df(ws, out, start_row=3, header_color=C["navy"], number_cols=num_cols)

    _set_col_widths(ws, {1: 28, 2: 11, 3: 14, 4: 11, 5: 14,
                         6: 11, 7: 14, 8: 11, 9: 14, 10: 11, 11: 14})
    _freeze(ws, "B4")


# ── 시트 3: 업종별 ────────────────────────────────────────────────────────────
def _sheet_category(wb: Workbook, df: pd.DataFrame, date_label: str, generated_at: str):
    ws = wb.create_sheet("업종별")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"업종별 결제 현황  ({date_label})"
    c.font = _font(bold=True, size=13, color=C["header_fg"])
    c.fill = _fill(C["dark"])
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H2")
    _meta_row(ws, 2, date_label, generated_at)

    out = df[["업종코드", "업종명", "가맹점_수", "전체_주문", "전체_결제",
              "카드_결제", "식권_결제", "쿠폰_결제", "승차권_결제", "비중(%)"]].copy()
    num_cols = ["가맹점_수", "전체_주문", "전체_결제", "카드_결제", "식권_결제", "쿠폰_결제", "승차권_결제"]
    _write_df(ws, out, start_row=3, header_color=C["blue"],
              number_cols=num_cols, pct_cols=["비중(%)"])

    _set_col_widths(ws, {1: 8, 2: 18, 3: 10, 4: 11, 5: 14,
                         6: 14, 7: 14, 8: 14, 9: 14, 10: 9})
    _freeze(ws, "C4")


# ── 시트 4: 가맹점별 ──────────────────────────────────────────────────────────
def _sheet_merchant(wb: Workbook, df: pd.DataFrame, date_label: str, generated_at: str):
    ws = wb.create_sheet("가맹점별")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = f"가맹점별 결제 현황  ({date_label})"
    c.font = _font(bold=True, size=13, color=C["header_fg"])
    c.fill = _fill(C["dark"])
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:K2")
    _meta_row(ws, 2, date_label, generated_at)

    out = df[["캠퍼스", "가맹점", "업종코드", "업종명",
              "전체_주문", "전체_결제", "카드_주문", "카드_결제",
              "식권_주문", "식권_결제", "쿠폰_주문", "쿠폰_결제",
              "승차권_주문", "승차권_결제"]].copy()
    num_cols = [c for c in out.columns if c not in ("캠퍼스", "가맹점", "업종코드", "업종명")]
    _write_df(ws, out, start_row=3, header_color=C["green"], number_cols=num_cols)

    _set_col_widths(ws, {1: 24, 2: 30, 3: 8, 4: 16,
                         5: 11, 6: 14, 7: 11, 8: 14,
                         9: 11, 10: 14, 11: 11, 12: 14,
                         13: 11, 14: 14})
    _freeze(ws, "C4")


# ── 시트 5: 미분류 알림 ───────────────────────────────────────────────────────
def _sheet_unmapped(wb: Workbook, df: pd.DataFrame, date_label: str, generated_at: str):
    ws = wb.create_sheet("미분류_알림")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    has = not df.empty
    c.value = (
        f"⚠  미분류 가맹점 알림  ({date_label})  —  총 {len(df)}개 업종 미지정"
        if has else
        f"✓  미분류 가맹점 없음  ({date_label})"
    )
    c.font = _font(bold=True, size=12, color=C["header_fg"])
    c.fill = _fill(C["warn"] if has else C["green"])
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:D2")
    c2 = ws["A2"]
    c2.value = "merchant_category.csv 또는 keyword_rules.csv 에 아래 가맹점을 추가하면 자동 분류됩니다."
    c2.font = _font(size=9, color=C["gray"])

    if has:
        _write_df(ws, df, start_row=3, header_color=C["warn"],
                  number_cols=["캠퍼스_수", "전체_결제"])

    _set_col_widths(ws, {1: 40, 2: 12, 3: 16})
    _freeze(ws, "A4")


# ── 공개 API ──────────────────────────────────────────────────────────────────
def build(
    *,
    output_path: Path,
    kpi: dict,
    df_campus: pd.DataFrame,
    df_category: pd.DataFrame,
    df_merchant: pd.DataFrame,
    df_unmapped: pd.DataFrame,
    date_label: str,
) -> Path:
    wb = Workbook()
    # 기본 시트 제거
    del wb[wb.sheetnames[0]]

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    _sheet_summary(wb, kpi, df_campus, df_category, date_label, generated_at)
    _sheet_campus(wb, df_campus, date_label, generated_at)
    _sheet_category(wb, df_category, date_label, generated_at)
    _sheet_merchant(wb, df_merchant, date_label, generated_at)
    _sheet_unmapped(wb, df_unmapped, date_label, generated_at)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
